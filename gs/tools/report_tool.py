import email
import json
import os
import datetime
import poplib
import re
import shutil
import time
from email.header import decode_header
from email.parser import Parser

import openpyxl
import logging

from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

logger = logging.getLogger(__name__)
logger.setLevel(level=logging.INFO)
handler = logging.FileHandler('./data/tools.log.%s' % datetime.datetime.now().year, mode='a+')
handler.setLevel(logging.INFO)
formatter = logging.Formatter('%(asctime)s - line:%(lineno)d - %(levelname)s: %(message)s')
handler.setFormatter(formatter)

console = logging.StreamHandler()
console.setLevel(logging.INFO)

logger.addHandler(handler)
logger.addHandler(console)

logger.info("开始运行")

class Qmail(object):
	def __init__(self,user,password):
		self.user = user  # 邮箱用户名
		self.password = password  # 邮箱密码
		self.server = None  # 初始化server，调用login_email方法后更新server
		self.mails = []  # 初始化邮箱信息列表，调用get_email_lists方法后更新列表

	def login_email(self):
		# 登录邮箱
		logger.info('登录QQ邮箱')
		pop3_server = 'pop.qq.com'
		try_times = 1
		while try_times<=3:
			try:
				server = poplib.POP3_SSL(host=pop3_server, port=995, timeout=50)
				# 身份认证:
				server.user(self.user)
				server.pass_(self.password)
				self.server = server
				logger.info(f'{self.user} 登录邮箱成功')
				return True
			except BaseException as e:
				try_times+=1
				logger.exception(e)
				logger.error(f'登录失败，尝试第{try_times}次登录')
				time.sleep(2)
		logger.error(f'3次登录失败！请检查用户名密码')
		return False

	def download_mails(self):
		# 获取邮件列表
		resp, mails, octets = self.server.list()  # list()返回所有邮件的编号:
		# 解析邮件
		for index in range(len(mails), 0, -1):
			self.mails.append(self.parser_mail(index))

	# 解析邮件
	def parser_mail(self, index):
		'''
		:param index: 邮件索引
		:return: 邮件正文、时间、主题、发件人的字典
		'''

		# 1、获取邮件原文
		resp, lines, octets = self.server.retr(index)  # 获取第index封邮件，lines存储了邮件的原始文本的每一行
		# 2、拼接邮件
		try:
			msg_content = b'\n'.join(lines).decode('gbk')  # 邮件的原始文本
		except:
			try:
				msg_content = b'\n'.join(lines).decode('utf-8')  # 邮件的原始文本
			except:
				return False

		# 3、解析邮件内容
		try:
			msg = Parser().parsestr(msg_content)
		except:
			return False

		# 4、解析邮件主题(标题)
		try:
			Subject = self.decode_str(msg.get("Subject"))
		except BaseException as e:
			return False

		# 5、解析邮件时间
		try:
			Date = time.strptime(self.decode_str(msg.get("Date"))[0:24], '%a, %d %b %Y %H:%M:%S')
			Date = time.mktime(Date)  # 获取邮件的接收时间,格式化收件时间
		except:
			return False

		# 6、解析发件人
		try:
			From = self.decode_str(msg.get("From")).split(' ')[-1]
		except:
			From = '<None>'

		return {
			'From': From,
			'Date': Date,
			'Subject': Subject,
			'Msg': msg,
		}
	# 字符编码转换
	def decode_str(self, str_in):
		try:
			value, charset = decode_header(str_in)[0]
			if charset:
				value = value.decode(charset)
			return value
		except:
			return str_in
	# 解析邮件,获取附件
	def get_att(self, msg_in,folder):
		attachment_files = []
		i = 1
		for part in msg_in.walk():
			# 获取附件名称类型
			file_name = part.get_filename()
			# contType = part.get_content_type()
			if file_name:
				h = email.header.Header(file_name)

				# 对附件名称进行解码
				dh = email.header.decode_header(h)
				filename = dh[0][0]
				if dh[0][1]:
					# 将附件名称可读化
					filename = self.decode_str(str(filename, dh[0][1]))
					# print(filename)
					# filename = filename.encode("utf-8")

				# 下载附件
				data = part.get_payload(decode=True)
				if not os.path.exists(folder):
					os.makedirs(folder)
				att_file = open(folder + '/' + filename, 'wb') # 注意二进制文件需要用wb模式打开
				attachment_files.append(filename)
				att_file.write(data)  # 保存附件
				att_file.close()

				logger.info(f'附件({i}): {filename}')
				i += 1
		return attachment_files
def load_config():
	with open('./data/config.txt', 'r', encoding='utf8') as f:
		data = f.readlines()
		data = [item.strip() for item in data if item]
		config = {}
		config['date'] = data[1].split('-')
		group = []
		mids = []
		small_count = 0
		for item in data[3:]:
			if not item:
				continue
			if item=='mail':
				continue
			if item.startswith('user='):
				config['user'] = item[5:]
				continue
			if item.startswith('password='):
				config['password'] = item[9:]
				continue
			mid, small, name = item.split('-')
			if mid not in mids:
				group.append({'name': mid, 'child': []})
				mids.append(mid)
			child = group[-1]['child']
			child.append({'name': small, 'ch': name})
			small_count += 1
		config['group'] = group
		return config, small_count

def check_small_excel_path(folder,data,date):
	small_excels = os.listdir(folder)
	small_excels_file_dict = {item: item.split('.')[0].split('-') for item in small_excels}
	for key, value in small_excels_file_dict.items():
		if data['name'] == value[0] and list(map(int, date)) == list(map(int, value[3:6])):
			return os.path.join(folder, key)
	return None

def new_run():
	config, small_count = load_config()
	logger.info('加载config配置\n{}'.format(config))
	if not config.get('user') or not config.get('password'):
		logger.warning('没有配置用户名密码，将无法登陆QQ邮箱')
		qmail = None
	else:
		qmail = Qmail(config['user'],config['password'])
		if qmail.login_email():
			qmail.download_mails()
		else:
			go = input('登陆邮箱失败，请问是否继续：（回车或者输入‘Y’继续）')
			if go in ['','Y','y']:
				pass
			else:
				logger.info('再见！')
				input()
				exit()
	date = config.get('date')
	group = config.get('group')
	flist = os.listdir('./')
	old_name_list = []
	for fname in flist:
		if fname.startswith('统计表-') and fname.endswith('.xlsx'):
			old_name_list.append(fname)
	if len(old_name_list)==1:
		old_name = old_name_list[0]
	else:
		t_old_name_list = [item.split('.')[0].split('-') for item in old_name_list]
		t_old_name_list.sort(key=lambda x:(int(x[1]),int(x[2]),int(x[3])))
		old_name = t_old_name_list[0]
		for item in t_old_name_list:
			if datetime.datetime(year=int(item[1]),month=int(item[2]),day=int(item[3]))>datetime.datetime(year=int(date[0]),month=int(date[1]),day=int(date[2])):
				break
			old_name= item
		t_old_name = '-'.join(old_name)
		for item in old_name_list:
			if item.startswith(t_old_name):
				old_name = item
				break
	logger.info('旧统计表：%s' % old_name)
	export_name = '统计表-{}-{}-{}.xlsx'.format(date[0], date[1], date[2])
	folder = './{}-{}'.format(date[0], date[1])
	if not os.path.exists(folder):
		logger.warning(f'日期文件夹【{folder}】不存在，将创建')
		os.mkdir(folder)
	small_excels = os.listdir(folder)
	logger.info('{}文件夹下的报表列表：\n{}'.format(folder,small_excels))
	small_excels_file_dict = {item: item.split('.')[0].split('-') for item in small_excels}
	# load 统计表
	wb_total = openpyxl.load_workbook(old_name)
	sheet_total = wb_total.worksheets[0]
	start_index = index = sheet_total.max_row + 1

	# style define
	font_bold = Font('微软雅黑', size=10, bold=True)
	align = Alignment(horizontal='center', vertical='center', wrap_text=False)
	font_normal = Font('微软雅黑', size=11, bold=False)
	fill_blue = PatternFill(fill_type='solid', start_color='00FDFF')
	side = Side(border_style='thin', color='000000')
	border = Border(left=side, right=side, top=side, bottom=side)
	head_fill_dict = {}
	for i in range(1, 19):
		head_fill_dict[i] = PatternFill(fill_type='solid', start_color='00FDFF')
	head_fill_dict[19] = PatternFill(fill_type='solid', start_color='FF99CC')
	head_fill_dict[20] = PatternFill(fill_type='solid', start_color='FF00FF')
	head_fill_dict[21] = PatternFill(fill_type='solid', start_color='FF8080')
	head_fill_dict[22] = PatternFill(fill_type='solid', start_color='FF0000')
	head_fill_dict[23] = PatternFill(fill_type='solid', start_color='FF99CC')
	head_fill_dict[24] = PatternFill(fill_type='solid', start_color='FF00FF')
	head_fill_dict[25] = PatternFill(fill_type='solid', start_color='FF8080')
	head_fill_dict[26] = PatternFill(fill_type='solid', start_color='FF0000')
	head_fill_dict[27] = PatternFill(fill_type='solid', start_color='FF9900')
	# format
	titles = ['Sunday', 'Large', 'Medium', 'Small', 'SF', 'L3+', 'L2', 'L1', 'L0', 'Adults', 'Children', 'Attendance',
			  'Newcomers', 'Absence', 'Care', 'Lost Sheep', 'Cover','Delete', 'Absen 1', 'Absen 2', 'Absen 3', 'Absen 4',
			  'Absen 1', 'Absen 2', 'Absen 3', 'Absen 4', 'Abs Ttl']
	sheet_total.append(titles)
	for i in range(1, 5):
		sheet_total.cell(index, i).font = font_bold
		sheet_total.cell(index, i).alignment = align
	for i in range(5, 28):
		sheet_total.cell(index, i).font = font_normal
		sheet_total.cell(index, i).alignment = align
	for i in range(1, 28):
		sheet_total.cell(index, i).border = border
		sheet_total.cell(index, i).fill = head_fill_dict[i]

	sum_list = []
	xiao_zong_index=[]
	for i in range(len(group)):
		mid_list = []
		for j in range(len(group[i]['child'])):
			index += 1
			data = group[i]['child'][j]
			logger.info('\n处理{}小家数据，data数据：{}'.format(data['name'],data))
			row_data = [None, None, None, data['ch']]
			small_excel_path = check_small_excel_path(folder,data,date)

			if not small_excel_path:
				logger.warning(f'{data["name"]}小家的报表文件不存在')
				if qmail:
					subject = f"{group[i]['name']}-{data['name']}-{date[0]}-{date[1]}-{date[2]}"
					logger.info(f"从邮箱下载{data['name']}小家的报表,邮件主题为：{subject}")
					find = False
					for mail_msg in qmail.mails:
						if mail_msg:
							p = re.match(r'.*({}'.format(f"{group[i]['name']}-{data['name']}")+r'-(\d\d\d\d)-(\d{1,2})-(\d{1,2})).*',mail_msg.get('Subject'))
							if p:
								groups = p.groups()
								if int(date[1]) == int(groups[2]) and int(date[2]) == int(groups[3]) and int(date[0]) == int(groups[1]):
									print(mail_msg.get('Subject'))
									attachment_files = qmail.get_att(mail_msg['Msg'],folder)  # 下载邮件中的附件
									small_excel_path = check_small_excel_path(folder, data, date)
									if not small_excel_path:
										if not attachment_files:
											logger.warning('邮件不含附件，继续搜索邮件')
											continue
										logger.warning('下载的附件中没有找到周报表，继续搜索邮件')
										continue
									find = True
									break
					if not find:
						logger.warning(f"未找到主题为【{subject}】的邮件")
					if small_excel_path:
						logger.info(f"{data['name']}小家的报表文件下载成功")
			if not small_excel_path:
				logger.warning(f'{data["name"]}小家的报表文件还是不存在')
				row_data.append(0)
				# color
				sheet_total.append(row_data)
				# font,align,border
				for a in range(1, 28):
					sheet_total.cell(index, a).font = font_normal
					sheet_total.cell(index, a).alignment = align
					sheet_total.cell(index, a).border = border
				mid_list.append(row_data)
				continue
			else:
				logger.warning(f'{data["name"]}小家的报表文件存在')
				row_data.append(1)
				ws_small = openpyxl.load_workbook(small_excel_path, read_only=True, data_only=True).worksheets[0]
				for a in range(1, ws_small.max_row):
					if ws_small.cell(a, 1).value == '人数统计' and ws_small.cell(a + 1, 1).value == '类别' and ws_small.cell(
							a + 2, 1).value == '颜色定义':
						index_small = a
						break
				adult_ab_list = []
				child_ab_list = []
				for a in range(7, 11):
					adult_ab_list.append(int(ws_small.cell(index_small + 3, a).value))
					child_ab_list.append(int(ws_small.cell(index_small + 4, a).value))
				attence_list = []
				for a in range(2, 15):
					attence_list.append(int(ws_small.cell(index_small + 9, a).value))
				row_data.extend(attence_list)
				row_data.extend(adult_ab_list)
				row_data.extend(child_ab_list)
				row_data.append(sum(child_ab_list) + sum(adult_ab_list))
				mid_list.append(row_data)
				sheet_total.append(row_data)
				# font,align,border
				for a in range(1, 28):
					sheet_total.cell(index, a).font = font_normal
					sheet_total.cell(index, a).alignment = align
					sheet_total.cell(index, a).border = border
				# color
				for a in range(19, 27):
					if sheet_total.cell(index, a).value:
						sheet_total.cell(index, a).fill = head_fill_dict[a]
						sheet_total.cell(index, 4).fill = head_fill_dict[a]

		mid_data = [None, None, group[i]['name'], '小总']
		for a in range(4, 27):
			sums = 0
			for item in mid_list:
				if len(item) <= a:
					continue
				sums += item[a]
			mid_data.append(sums)
		index += 1
		xiao_zong_index.append(index)
		# merge_cell
		sheet_total.merge_cells(start_row=index - len(mid_list), start_column=3, end_row=index, end_column=3)
		sheet_total.cell(index - len(mid_list), 3, mid_data[2])
		sheet_total.cell(index,4,mid_data[3])
		col_list=['E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z','AA']
		for col in col_list:
			sheet_total['{}{}'.format(col,index)].value = '=SUM({}{}:{}{})'.format(col,index-len(mid_list),col,index-1)
		# font,align,border
		for a in range(1, 28):
			sheet_total.cell(index, a).font = font_normal
			sheet_total.cell(index, a).alignment = align
			sheet_total.cell(index, a).border = border
		for a in range(4, 28):
			sheet_total.cell(index, a).fill = fill_blue
		sum_list.append(mid_data)
	sum_data = ["{}月{}日".format(date[1], date[2]), 'HOD', "HOD", 'Total']
	for a in range(4, 27):
		sums = 0
		for item in sum_list:
			sums += item[a]
		sum_data.append(sums)
	index += 1
	# merge_cell·
	sheet_total.merge_cells(start_row=start_index + 1, start_column=1, end_row=index, end_column=1)
	sheet_total.merge_cells(start_row=start_index + 1, start_column=2, end_row=index, end_column=2)
	sheet_total.cell(start_index + 1, 1, sum_data[0])
	sheet_total.cell(start_index + 1, 2, sum_data[1])
	sheet_total.cell(index,3,sum_data[2])
	sheet_total.cell(index,4,sum_data[3])
	col_list = ['E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y',
				'Z','AA']
	for col in col_list:
		sheet_total['{}{}'.format(col,index)].value = '='+'+'.join([col+str(t_index) for t_index in xiao_zong_index])
	# font,align,border
	for a in range(1, 28):
		sheet_total.cell(index, a).font = font_normal
		sheet_total.cell(index, a).alignment = align
		sheet_total.cell(index, a).border = border
	for a in range(4, 28):
		sheet_total.cell(index, a).fill = PatternFill(fill_type='solid', start_color='FFFF00')
	shutil.copyfile(old_name, './data/backup.xlsx')
	wb_total.save(export_name)
	logger.info('恭喜，生成报表成功！报表名：{}\n\n'.format(export_name))


if __name__ == '__main__':
	try:
		new_run()
	except Exception as e:
		logger.exception(e)
		logger.error('！！！！出错了，请查看上面错误日志或者联系开发人员')
	input('请按任意键结束：')
