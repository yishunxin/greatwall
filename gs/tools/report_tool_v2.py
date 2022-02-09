import email
import json
import os
import datetime
import imaplib
import poplib

import re
import shutil
import time
from email.header import decode_header
from email.parser import BytesParser
from email.parser import Parser
import win32com.client as wind32

import openpyxl
import logging

from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

logger = logging.getLogger(__name__)
logger.setLevel(level=logging.INFO)
handler = logging.FileHandler('./data/tools.log.%s' % datetime.datetime.now().year, mode='a+')
handler.setLevel(logging.INFO)
formatter = logging.Formatter('%(asctime)s - line:%(lineno)d - %(levelname)s: %(message)s')
handler.setFormatter(formatter)

console = logging.StreamHandler()
console.setLevel(logging.INFO)

logger.addHandler(handler)
logger.addHandler(console)

logger.info("开始运行")


class ImapMail(object):
	def __init__(self, user, password):
		self.imap_host = 'imap.qq.com'
		self.user = user
		self.password = password
		self.server = None
		self.mail_list = []

	def login_email(self):
		try_times = 1
		logger.info('登录QQ邮箱')
		while try_times <= 3:
			try:
				self.server = imaplib.IMAP4(self.imap_host)
				self.server.login(self.user, self.password)
				logger.info(f'{self.user} 登录邮箱成功')
				self.server.select()
				return True
			except Exception as e:
				try_times += 1
				logger.exception(e)
				logger.error(f'登录失败，尝试第{try_times}次登录')
				time.sleep(2)
		logger.error(f'3次登录失败！请检查用户名密码')
		return False

	def decode_str(self, s):
		try:
			subject = email.header.decode_header(s)
		except Exception as e:
			logger.exception(e)
			return None
		sub_bytes = subject[0][0]
		sub_charset = subject[0][1]
		if None == sub_charset:
			subject = sub_bytes
		elif 'unknown-8bit' == sub_charset:
			subject = str(sub_bytes, 'utf8')
		else:
			subject = str(sub_bytes, sub_charset)
		return subject

	def scan_mail_has_file(self):
		logger.info('开始扫描邮件,耗时1分钟以内，请稍等')
		status, data = self.server.search(None, 'ALL')
		email_list = list(reversed(data[0].split()))
		for num in email_list:
			typ, content = self.server.fetch(num, '(RFC822)')
			msg = BytesParser().parsebytes(content[0][1])
			for part in msg.walk():
				fileName = part.get_filename()
				if fileName is not None:
					fileName = self.decode_str(fileName)
					self.mail_list.append([fileName, msg, part])

	def download_file(self, folder, filename, part):
		try:
			logger.info('下载附件：%s' % filename)
			if not os.path.exists(folder):
				os.makedirs(folder)
			with open(os.path.join(folder, filename), 'wb') as f:
				f.write(part.get_payload(decode=True))
			return True
		except Exception as e:
			logger.exception(e)
			return False

	def logout(self):
		if self.server:
			self.server.close()
			self.server.logout()


def download_all_excel(config):
	group = config.get('group')
	date = config.get('date')
	folder = './{}-{}'.format(date[0], date[1])
	fail_list = []
	logger.info('开始下载各小家ZBB')
	qmail = login_mail(config.get('user'), config.get('password'))
	for i in range(len(group)):
		for j in range(len(group[i]['child'])):
			data = group[i]['child'][j]
			small_excel_path = check_small_excel_path(folder, group[i]['name'], data['name'], date)
			if small_excel_path:
				if data.get('download_flag', False):
					logger.info(f'{data["name"]}小家的报表文件存在,无需下载')
					data['download_flag'] = True
				continue
			logger.warning(f'{data["name"]}小家的报表文件不存在')
			if qmail:
				subject = f"{group[i]['name']}-{data['name']}-{date[0]}-{date[1]}-{date[2]}"
				logger.info(f"从邮箱下载{data['name']}小家的报表,邮件主题为：{subject}")
				for filename, mail_msg, part in qmail.mail_list:
					try:
						if mail_msg:
							mail_subject = qmail.decode_str(mail_msg.get("Subject"))

							new_filename = filename.replace(' ', '')
							m = re.match(
								r'([a-zA-Z]+)[-|—|_]+([a-zA-Z]+)[-|—|_]+(\d+)[-|—|_]+(\d+)[-|—|_]+(\d+).([a-zA-Z]+)',
								new_filename)
							if not m:
								continue
							key_list = m.groups()
							if group[i]['name'] == key_list[0].upper() and data['name'] == key_list[
								1].upper() and list(map(int, date[0:3])) == list(map(int, key_list[2:5])):
								print(mail_subject)
								logger.info(f'找到报表附件{filename}')
								should_name = f"{group[i]['name']}-{data['name']}-{date[0]}-{date[1]}-{date[2]}.{key_list[-1]}"
								if filename != should_name:
									logger.warning(f'报表附件名【{filename}】不规范，将保存为【{should_name}】')
								bo = qmail.download_file(folder, should_name, part)
								if not bo:
									logger.error('下载附件失败')
									continue
								logger.info('下载附件成功')
								small_excel_path = check_small_excel_path(folder, group[i]['name'], data['name'], date)
								if not small_excel_path:
									logger.warning('下载的附件中没有找到周报表，继续搜索邮件')
									continue
								else:
									logger.info(f"{data['name']}小家的报表文件下载成功")
								break
					except Exception as e:
						logger.exception(e)
						logger.error(f"从邮箱下载{data['name']}小家报表文件出错，先跳过")
			small_excel_path = check_small_excel_path(folder, group[i]['name'], data['name'], date)
			if not small_excel_path:
				fail_list.append(data["name"])
			else:
				data['download_flag'] = True
	if fail_list:
		redo = input("（{}）小家报表不存在，是否重新登录邮箱尝试下载：（回车或者输入‘Y’）".format("  ".join(fail_list)))
		if redo in ['', 'Y', 'y']:
			if qmail:
				qmail.logout()
			download_all_excel(config)
		else:
			pass


def login_mail(user, password):
	if not user or not password:
		logger.warning('没有配置用户名密码，将无法登陆QQ邮箱')
		qmail = None
	else:
		qmail = ImapMail(user, password)
		if qmail.login_email():
			qmail.scan_mail_has_file()
		else:
			go = input('登陆邮箱失败，请问是否继续：（回车或者输入‘Y’继续）')
			if go in ['', 'Y', 'y']:
				pass
			else:
				logger.info('再见！')
				input()
				exit()
	return qmail


def run():
	config, small_count = load_config()
	logger.info('加载config配置\n{}'.format(config))
	date = config.get('date')
	group = config.get('group')
	flist = os.listdir('./')
	old_name_list = []
	for fname in flist:
		if fname.startswith('统计表-') and fname.endswith('.xlsx'):
			old_name_list.append(fname)
	if len(old_name_list) == 1:
		old_name = old_name_list[0]
	else:
		t_old_name_list = [item.split('.')[0].split('-') for item in old_name_list]
		t_old_name_list.sort(key=lambda x: (int(x[1]), int(x[2]), int(x[3])))
		old_name = t_old_name_list[0]
		for item in t_old_name_list:
			if datetime.datetime(year=int(item[1]), month=int(item[2]), day=int(item[3])) > datetime.datetime(
					year=int(date[0]), month=int(date[1]), day=int(date[2])):
				break
			old_name = item
		t_old_name = '-'.join(old_name)
		for item in old_name_list:
			if item.startswith(t_old_name):
				old_name = item
				break
	logger.info('旧统计表：%s' % old_name)
	export_name = '统计表-{}-{}-{}.xlsx'.format(date[0], date[1], date[2])
	folder = './{}-{}'.format(date[0], date[1])
	if not os.path.exists(folder):
		logger.warning(f'日期文件夹【{folder}】不存在，将创建')
		os.mkdir(folder)
	small_excels = os.listdir(folder)
	logger.info('{}文件夹下的报表列表：\n{}'.format(folder, small_excels))
	small_excels_file_dict = {item: item.split('.')[0].split('-') for item in small_excels}
	# load 统计表
	wb_total = openpyxl.load_workbook(old_name)
	sheet_total = wb_total.worksheets[0]
	start_index = index = sheet_total.max_row + 1

	# style define
	font_bold = Font('微软雅黑', size=10, bold=True)
	align = Alignment(horizontal='center', vertical='center', wrap_text=False)
	font_normal = Font('微软雅黑', size=11, bold=False)
	fill_blue = PatternFill(fill_type='solid', start_color='00FDFF')
	side = Side(border_style='thin', color='000000')
	border = Border(left=side, right=side, top=side, bottom=side)
	head_fill_dict = {}
	for i in range(1, 19):
		head_fill_dict[i] = PatternFill(fill_type='solid', start_color='00FDFF')
	head_fill_dict[19] = PatternFill(fill_type='solid', start_color='FF99CC')
	head_fill_dict[20] = PatternFill(fill_type='solid', start_color='FF00FF')
	head_fill_dict[21] = PatternFill(fill_type='solid', start_color='FF8080')
	head_fill_dict[22] = PatternFill(fill_type='solid', start_color='FF0000')
	head_fill_dict[23] = PatternFill(fill_type='solid', start_color='FF99CC')
	head_fill_dict[24] = PatternFill(fill_type='solid', start_color='FF00FF')
	head_fill_dict[25] = PatternFill(fill_type='solid', start_color='FF8080')
	head_fill_dict[26] = PatternFill(fill_type='solid', start_color='FF0000')
	head_fill_dict[27] = PatternFill(fill_type='solid', start_color='FF9900')
	# format
	titles = ['Sunday', 'Large', 'Medium', 'Small', 'SF', 'L3+', 'L2', 'L1', 'L0', 'Adults', 'Children', 'Attendance',
			  'Newcomers', 'Absence', 'Care', 'Lost Sheep', 'Cover', 'Delete', 'Absen 1', 'Absen 2', 'Absen 3',
			  'Absen 4',
			  'Absen 1', 'Absen 2', 'Absen 3', 'Absen 4', 'Abs Ttl']
	sheet_total.append(titles)
	for i in range(1, 5):
		sheet_total.cell(index, i).font = font_bold
		sheet_total.cell(index, i).alignment = align
	for i in range(5, 28):
		sheet_total.cell(index, i).font = font_normal
		sheet_total.cell(index, i).alignment = align
	for i in range(1, 28):
		sheet_total.cell(index, i).border = border
		sheet_total.cell(index, i).fill = head_fill_dict[i]

	sum_list = []
	xiao_zong_index = []
	download_all_excel(config)
	for i in range(len(group)):
		mid_list = []
		for j in range(len(group[i]['child'])):
			index += 1
			data = group[i]['child'][j]
			logger.info('\n处理{}小家数据，data数据：{}'.format(data['name'], data))
			row_data = [None, None, None, data['ch']]
			small_excel_path = check_small_excel_path(folder, group[i]['name'], data['name'], date)
			if not small_excel_path:
				logger.warning(f'{data["name"]}小家的报表文件还是不存在')
				row_data.append(0)
				# color
				sheet_total.append(row_data)
				# font,align,border
				for a in range(1, 28):
					sheet_total.cell(index, a).font = font_normal
					sheet_total.cell(index, a).alignment = align
					sheet_total.cell(index, a).border = border
				mid_list.append(row_data)
				continue
			else:
				logger.info(f'{data["name"]}小家的报表文件存在')
				row_data.append(1)
				if small_excel_path.split('.')[-1] == 'xls':
					logger.info(f'{data["name"]}小家的报表文件为xls格式，执行格式转换。（请确保电脑装有较新的excel软件）')
					try:
						xls2xlsx(os.path.join(os.getcwd(), small_excel_path))
						small_excel_path += 'x'
						logger.info("格式转换成功")
					except Exception as e:
						logger.exception(e)
						raise RuntimeError("格式转换失败")
				ws_small = openpyxl.load_workbook(small_excel_path, read_only=True, data_only=True).worksheets[0]
				for a in range(1, ws_small.max_row):
					if ws_small.cell(a, 1).value == '人数统计' and ws_small.cell(a + 1, 1).value == '类别' and ws_small.cell(
							a + 2, 1).value == '颜色定义':
						index_small = a
						break
				adult_ab_list = []
				child_ab_list = []
				for a in range(7, 11):
					adult_ab_list.append(int(ws_small.cell(index_small + 3, a).value))
					child_ab_list.append(int(ws_small.cell(index_small + 4, a).value))
				attence_list = []
				for a in range(2, 15):
					attence_list.append(int(ws_small.cell(index_small + 9, a).value))
				row_data.extend(attence_list)
				row_data.extend(adult_ab_list)
				row_data.extend(child_ab_list)
				row_data.append(sum(child_ab_list) + sum(adult_ab_list))
				mid_list.append(row_data)
				sheet_total.append(row_data)
				# font,align,border
				for a in range(1, 28):
					sheet_total.cell(index, a).font = font_normal
					sheet_total.cell(index, a).alignment = align
					sheet_total.cell(index, a).border = border
				# color
				for a in range(19, 27):
					if sheet_total.cell(index, a).value:
						sheet_total.cell(index, a).fill = head_fill_dict[a]
						sheet_total.cell(index, 4).fill = head_fill_dict[a]

		mid_data = [None, None, group[i]['name'], '小总']
		for a in range(4, 27):
			sums = 0
			for item in mid_list:
				if len(item) <= a:
					continue
				sums += item[a]
			mid_data.append(sums)
		index += 1
		xiao_zong_index.append(index)
		# merge_cell
		sheet_total.merge_cells(start_row=index - len(mid_list), start_column=3, end_row=index, end_column=3)
		sheet_total.cell(index - len(mid_list), 3, mid_data[2])
		sheet_total.cell(index, 4, mid_data[3])
		col_list = ['E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X',
					'Y', 'Z', 'AA']
		for col in col_list:
			sheet_total['{}{}'.format(col, index)].value = '=SUM({}{}:{}{})'.format(col, index - len(mid_list), col,
																					index - 1)
		# font,align,border
		for a in range(1, 28):
			sheet_total.cell(index, a).font = font_normal
			sheet_total.cell(index, a).alignment = align
			sheet_total.cell(index, a).border = border
		for a in range(4, 28):
			sheet_total.cell(index, a).fill = fill_blue
		sum_list.append(mid_data)
	sum_data = ["{}月{}日".format(date[1], date[2]), 'HOD', "HOD", 'Total']
	for a in range(4, 27):
		sums = 0
		for item in sum_list:
			sums += item[a]
		sum_data.append(sums)
	index += 1
	# merge_cell·
	sheet_total.merge_cells(start_row=start_index + 1, start_column=1, end_row=index, end_column=1)
	sheet_total.merge_cells(start_row=start_index + 1, start_column=2, end_row=index, end_column=2)
	sheet_total.cell(start_index + 1, 1, sum_data[0])
	sheet_total.cell(start_index + 1, 2, sum_data[1])
	sheet_total.cell(index, 3, sum_data[2])
	sheet_total.cell(index, 4, sum_data[3])
	col_list = ['E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y',
				'Z', 'AA']
	for col in col_list:
		sheet_total['{}{}'.format(col, index)].value = '=' + '+'.join(
			[col + str(t_index) for t_index in xiao_zong_index])
	# font,align,border
	for a in range(1, 28):
		sheet_total.cell(index, a).font = font_normal
		sheet_total.cell(index, a).alignment = align
		sheet_total.cell(index, a).border = border
	for a in range(4, 28):
		sheet_total.cell(index, a).fill = PatternFill(fill_type='solid', start_color='FFFF00')
	shutil.copyfile(old_name, './data/backup.xlsx')
	wb_total.save(export_name)
	logger.info('恭喜，生成报表成功！报表名：{}\n\n'.format(export_name))


def load_config():
	with open('./data/config.txt', 'r', encoding='utf8') as f:
		data = f.readlines()
		data = [item.strip() for item in data if item]
		config = {}
		config['date'] = data[1].split('-')
		group = []
		mids = []
		small_count = 0
		for item in data[3:]:
			if not item:
				continue
			if item == 'mail':
				continue
			if item.startswith('user='):
				config['user'] = item[5:]
				continue
			if item.startswith('password='):
				config['password'] = item[9:]
				continue
			mid, small, name = item.split('-')
			if mid not in mids:
				group.append({'name': mid, 'child': []})
				mids.append(mid)
			child = group[-1]['child']
			child.append({'name': small, 'ch': name})
			small_count += 1
		config['group'] = group
		return config, small_count


def xls2xlsx(fname):
	excel = wind32.gencache.EnsureDispatch('Excel.Application')
	wb = excel.Workbooks.Open(fname)
	wb.SaveAs(fname + "x", FileFormat=51)
	wb.Close()
	excel.Application.Quit()


def check_small_excel_path(folder, big_name, small_name, date):
	small_excels = os.listdir(folder)
	small_excels_file_dict = {item: item.split('.')[0].split('-') for item in small_excels}
	posible_file = []
	for key, value in small_excels_file_dict.items():
		if big_name == value[0] and small_name == value[1] and list(map(int, date)) == list(map(int, value[2:5])):
			posible_file.append(key)
	if not posible_file:
		return None
	if len(posible_file) > 1:
		for item in posible_file:
			if item.endswith('.xlsx'):
				return os.path.join(folder, item)
	return os.path.join(folder, posible_file[0])


if __name__ == '__main__':
	try:
		run()
	except Exception as e:
		logger.exception(e)
		logger.error('！！！！出错了，请查看上面错误日志或者联系开发人员')
	input('请按任意键结束：')
