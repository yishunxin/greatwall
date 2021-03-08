
import openpyxl
from openpyxl.cell import MergedCell
from openpyxl.styles import PatternFill

w = openpyxl.load_workbook('123.xlsx')
ws = w.worksheets[0]
ws.unmerge_cells(start_row=1,start_column=1,end_row=1,end_column=30)
w.save('123.xlsx')
